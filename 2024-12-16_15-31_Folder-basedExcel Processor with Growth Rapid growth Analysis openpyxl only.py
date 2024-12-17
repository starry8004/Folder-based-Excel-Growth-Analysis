import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from datetime import datetime

class FolderExcelProcessor:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("셀링하니 상품 분석기")
        self.root.geometry("600x400")
        self.setup_gui()

    def setup_gui(self):
        # 메인 프레임
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # 상태 레이블
        self.status_label = ttk.Label(self.main_frame, text="폴더를 선택해주세요")
        self.status_label.pack(pady=10)

        # 진행바
        self.progress_frame = ttk.Frame(self.main_frame)
        self.progress_frame.pack(fill=tk.X, pady=10)
        
        # 전체 진행바
        self.total_progress_label = ttk.Label(self.progress_frame, text="전체 진행률:")
        self.total_progress_label.pack()
        self.total_progress = ttk.Progressbar(self.progress_frame, length=400, mode='determinate')
        self.total_progress.pack(pady=5)
        
        # 현재 파일 진행바
        self.file_progress_label = ttk.Label(self.progress_frame, text="현재 파일 진행률:")
        self.file_progress_label.pack()
        self.file_progress = ttk.Progressbar(self.progress_frame, length=400, mode='determinate')
        self.file_progress.pack(pady=5)

        # 현재 처리 중인 파일 레이블
        self.current_file_label = ttk.Label(self.main_frame, text="")
        self.current_file_label.pack(pady=5)

        # 폴더 선택 버튼
        self.select_button = ttk.Button(self.main_frame, text="폴더 선택", command=self.process_folder)
        self.select_button.pack(pady=20)

    def get_column_indices(self, sheet):
        headers = {}
        for idx, cell in enumerate(sheet[1], 1):
            headers[cell.value] = idx
        return headers

    def convert_to_float(self, value):
        if isinstance(value, (int, float)):
            return float(value)
        elif isinstance(value, str):
            try:
                return float(value.replace(',', ''))
            except ValueError:
                return 0.0
        return 0.0

    def process_single_file(self, filepath, analysis_type):
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # 컬럼 인덱스 찾기
        headers = self.get_column_indices(sheet)
        required_columns = ['성장성', '검색량', '쇼핑성키워드', '경쟁률', '키워드', 
                          '카테고리전체', '광고경쟁강도', '계절성']
        
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"필수 컬럼 '{col}'을 찾을 수 없습니다.")

        # 새 워크북 생성
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        
        # 헤더 작성
        new_headers = ['순서_검색량순', '키워드', '카테고리전체', '검색량', '경쟁률', 
                      '광고경쟁강도', '계절성']
        for col, header in enumerate(new_headers, 1):
            new_sheet.cell(1, col, header)

        # 데이터 필터링 및 정렬을 위한 임시 리스트
        filtered_data = []
        total_rows = sheet.max_row - 1

        for row in range(2, sheet.max_row + 1):
            self.file_progress['value'] = (row/total_rows * 100)
            self.root.update()

            growth = self.convert_to_float(sheet.cell(row, headers['성장성']).value)
            search_volume = self.convert_to_float(sheet.cell(row, headers['검색량']).value)
            is_shopping = str(sheet.cell(row, headers['쇼핑성키워드']).value).lower() == 'true'
            competition = self.convert_to_float(sheet.cell(row, headers['경쟁률']).value)

            meets_criteria = False
            if analysis_type == "growth":
                meets_criteria = (growth >= 0 and search_volume >= 8000 and 
                                is_shopping and competition < 4)
            else:  # rapid_growth
                meets_criteria = (growth >= 0.15 and search_volume >= 10000 and 
                                is_shopping)

            if meets_criteria:
                row_data = [
                    sheet.cell(row, headers['키워드']).value,
                    sheet.cell(row, headers['카테고리전체']).value,
                    search_volume,
                    competition,
                    sheet.cell(row, headers['광고경쟁강도']).value,
                    sheet.cell(row, headers['계절성']).value
                ]
                filtered_data.append(row_data)

        # 검색량 기준 정렬
        filtered_data.sort(key=lambda x: x[2], reverse=True)

        # 정렬된 데이터 쓰기
        for idx, row_data in enumerate(filtered_data, 1):
            new_sheet.cell(idx + 1, 1, idx)  # 순서_검색량순
            for col, value in enumerate(row_data, 2):
                new_sheet.cell(idx + 1, col, value)

        return new_wb, len(filtered_data), total_rows

    def process_folder(self):
        folder_path = filedialog.askdirectory(title="분석할 엑셀 파일이 있는 폴더를 선택하세요")
        if not folder_path:
            return

        try:
            # 엑셀 파일 목록 가져오기
            excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and '셀하' in f]
            if not excel_files:
                messagebox.showerror("에러", "처리할 엑셀 파일을 찾을 수 없습니다.")
                return

            total_files = len(excel_files) * 2  # 각 파일당 성장/급성장 2번 처리
            processed_files = 0

            for excel_file in excel_files:
                input_file = os.path.join(folder_path, excel_file)
                self.current_file_label.config(text=f"처리 중: {excel_file}")
                
                # 성장 분석
                try:
                    self.status_label.config(text=f"{excel_file} - 성장 상품 분석 중...")
                    new_wb, filtered_count, total_rows = self.process_single_file(input_file, "growth")
                    current_time = datetime.now().strftime('%Y-%m-%d_%H-%M')
                    output_file = os.path.join(folder_path, f"{os.path.splitext(excel_file)[0]}_성장_{current_time}.xlsx")
                    new_wb.save(output_file)
                    processed_files += 1
                    self.total_progress['value'] = (processed_files/total_files * 100)
                except Exception as e:
                    messagebox.showerror("에러", f"{excel_file} 성장 분석 중 오류 발생:\n{str(e)}")

                # 급성장 분석
                try:
                    self.status_label.config(text=f"{excel_file} - 급성장 상품 분석 중...")
                    new_wb, filtered_count, total_rows = self.process_single_file(input_file, "rapid_growth")
                    current_time = datetime.now().strftime('%Y-%m-%d_%H-%M')
                    output_file = os.path.join(folder_path, f"{os.path.splitext(excel_file)[0]}_급성장_{current_time}.xlsx")
                    new_wb.save(output_file)
                    processed_files += 1
                    self.total_progress['value'] = (processed_files/total_files * 100)
                except Exception as e:
                    messagebox.showerror("에러", f"{excel_file} 급성장 분석 중 오류 발생:\n{str(e)}")

            messagebox.showinfo("완료", 
                f"모든 파일 처리가 완료되었습니다!\n\n"
                f"처리된 파일 수: {len(excel_files)}개\n"
                f"저장 위치: {folder_path}")

        except Exception as e:
            messagebox.showerror("에러", f"처리 중 오류가 발생했습니다:\n{str(e)}")
        
        finally:
            self.total_progress['value'] = 0
            self.file_progress['value'] = 0
            self.current_file_label.config(text="")
            self.status_label.config(text="폴더를 선택해주세요")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = FolderExcelProcessor()
    app.run()